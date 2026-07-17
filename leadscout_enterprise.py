import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import re
import pandas as pd
import threading
from groq import Groq
from playwright.sync_api import sync_playwright
import urllib.parse
import time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 🔑 AKTİF AI SÜZGECİ ANAHTARIN
GROQ_API_KEY = "-----"

def _round_rect_points(x1, y1, x2, y2, radius):
    return [
        x1 + radius, y1,
        x2 - radius, y1,
        x2, y1,
        x2, y1 + radius,
        x2, y2 - radius,
        x2, y2,
        x2 - radius, y2,
        x1 + radius, y2,
        x1, y2,
        x1, y2 - radius,
        x1, y1 + radius,
        x1, y1,
    ]


class RoundedButton(tk.Canvas):
    """Gerçek yuvarlak köşeli, düz (flat) modern buton.
    Windows'un standart tk.Button'ının verdiği eski/kabarık (bevel) görünüm
    yerine, Canvas üzerine çizilen ince ve düz bir buton kullanır."""

    def __init__(self, parent, text, command, bg, fg, hover_bg=None,
                 disabled_bg="#334155", disabled_fg="#64748b",
                 width=260, height=44, radius=12,
                 font=("Segoe UI", 10, "bold"), parent_bg="#0b0f19"):
        super().__init__(parent, width=width, height=height, bg=parent_bg,
                          highlightthickness=0, bd=0)
        self.command = command
        self.bg_color = bg
        self.hover_bg = hover_bg or bg
        self.fg_color = fg
        self.disabled_bg = disabled_bg
        self.disabled_fg = disabled_fg
        self.disabled = False
        self.width = width
        self.height = height

        pts = _round_rect_points(1, 1, width - 1, height - 1, radius)
        self.shape = self.create_polygon(pts, fill=bg, outline="", smooth=True)
        self.label = self.create_text(width / 2, height / 2, text=text, fill=fg, font=font)

        self.bind("<Button-1>", self._on_click)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def _on_click(self, _event):
        if not self.disabled and self.command:
            self.command()

    def _on_enter(self, _event):
        if not self.disabled:
            self.itemconfig(self.shape, fill=self.hover_bg)
            self.config(cursor="hand2")

    def _on_leave(self, _event):
        if not self.disabled:
            self.itemconfig(self.shape, fill=self.bg_color)

    def set_state(self, state):
        if state == "disabled":
            self.disabled = True
            self.itemconfig(self.shape, fill=self.disabled_bg)
            self.itemconfig(self.label, fill=self.disabled_fg)
            self.config(cursor="arrow")
        else:
            self.disabled = False
            self.itemconfig(self.shape, fill=self.bg_color)
            self.itemconfig(self.label, fill=self.fg_color)

    def config(self, **kwargs):
        # tk.Button uyumluluğu için: .config(state="disabled"/"normal") çağrılarını yakala
        if "state" in kwargs and len(kwargs) == 1:
            self.set_state(kwargs["state"])
            return
        super().config(**kwargs)


class LeadScoutEnterprise:
    def __init__(self, root):
        self.root = root
        self.root.title("LeadScout Enterprise v3.5 🛰️")
        self.root.geometry("1020x720")
        self.root.minsize(880, 600)
        self.root.configure(bg="#0b0f19")
        
        self.leads_data = []
        if GROQ_API_KEY:
            self.groq_client = Groq(api_key=GROQ_API_KEY)
        else:
            self.groq_client = None

        self.create_ui()

    def create_ui(self):
        # ---- Renk Paleti ----
        BG = "#0b0f19"
        CARD = "#111827"
        BORDER = "#1f2937"
        TEXT = "#f1f5f9"
        MUTED = "#94a3b8"
        ACCENT = "#38bdf8"
        PRIMARY = "#6366f1"
        PRIMARY_HOVER = "#7c7ff2"
        SUCCESS = "#22c55e"
        SUCCESS_HOVER = "#34d774"
        FIELD_BG = "#0f1524"

        # Segoe UI Windows'ta her zaman kurulu gelir ve modern görünür.
        # Eskiden kullanılan "Inter" / "Plus Jakarta Sans" fontları sistemde
        # yoksa Tk sessizce eski/varsayılan bir fonta düşüyordu (XP görünümü
        # buradan kaynaklanıyordu).
        F_LABEL = ("Segoe UI", 9, "bold")
        F_ENTRY = ("Segoe UI", 10)
        F_TITLE = ("Segoe UI", 16, "bold")
        F_SUBTITLE = ("Segoe UI", 9)
        F_STATUS = ("Segoe UI", 10)

        self.root.configure(bg=BG)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                         background=CARD, foreground=TEXT, fieldbackground=CARD,
                         borderwidth=0, rowheight=30, font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
                         background="#1e293b", foreground=ACCENT,
                         font=("Segoe UI", 10, "bold"), relief="flat", borderwidth=0)
        style.map("Treeview.Heading", background=[("active", "#1e293b")])
        style.map("Treeview", background=[("selected", PRIMARY)], foreground=[("selected", "#ffffff")])
        style.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])

        # ================= HEADER =================
        header = tk.Frame(self.root, bg=BG)
        header.pack(fill="x", padx=24, pady=(20, 10))

        tk.Label(header, text="🛰️", bg=BG, font=("Segoe UI", 22)).pack(side="left", padx=(0, 10))
        title_box = tk.Frame(header, bg=BG)
        title_box.pack(side="left")
        tk.Label(title_box, text="LeadScout Enterprise", bg=BG, fg=TEXT, font=F_TITLE).pack(anchor="w")
        tk.Label(title_box, text="Google Haritalar Üzerinden Kurumsal Lead Madenciliği",
                 bg=BG, fg=MUTED, font=F_SUBTITLE).pack(anchor="w")

        badge = tk.Label(header, text="v3.5", bg="#1e293b", fg=ACCENT, font=("Segoe UI", 8, "bold"), padx=10, pady=4)
        badge.pack(side="right")

        # ================= FİLTRE KARTI =================
        card = tk.Frame(self.root, bg=CARD, highlightthickness=1, highlightbackground=BORDER, highlightcolor=BORDER)
        card.pack(fill="x", padx=24, pady=10)
        inner = tk.Frame(card, bg=CARD)
        inner.pack(fill="x", padx=20, pady=18)

        def field_label(parent, text, row, col):
            tk.Label(parent, text=text, bg=CARD, fg=ACCENT, font=F_LABEL).grid(
                row=row, column=col, padx=(0, 8), pady=10, sticky="w")

        def make_entry(parent, width=22):
            return tk.Entry(parent, bg=FIELD_BG, fg=TEXT, insertbackground=ACCENT,
                             relief="flat", bd=0, highlightthickness=1,
                             highlightbackground=BORDER, highlightcolor=PRIMARY,
                             font=F_ENTRY, width=width)

        inner.grid_columnconfigure(1, weight=1)
        inner.grid_columnconfigure(3, weight=1)

        field_label(inner, "Hedef Sektör", 0, 0)
        self.ent_keyword = make_entry(inner, width=24)
        self.ent_keyword.grid(row=0, column=1, padx=(0, 24), pady=10, sticky="ew", ipady=5)
        self.ent_keyword.insert(0, "Paslanmaz")

        field_label(inner, "Hedef Kayıt Sayısı", 0, 2)
        self.spin_limit = tk.Spinbox(inner, from_=10, to=1000, increment=10, width=8,
                                      bg=FIELD_BG, fg=TEXT, insertbackground=ACCENT,
                                      relief="flat", bd=0, highlightthickness=1,
                                      highlightbackground=BORDER, highlightcolor=PRIMARY,
                                      buttonbackground="#1e293b", font=F_ENTRY)
        self.spin_limit.grid(row=0, column=3, padx=(0, 0), pady=10, sticky="w", ipady=5)

        field_label(inner, "Hedef İl", 1, 0)
        self.ent_city = make_entry(inner, width=24)
        self.ent_city.grid(row=1, column=1, padx=(0, 24), pady=10, sticky="ew", ipady=5)
        self.ent_city.insert(0, "İstanbul")

        field_label(inner, "Hedef İlçe (Opsiyonel)", 1, 2)
        self.ent_district = make_entry(inner, width=16)
        self.ent_district.grid(row=1, column=3, padx=(0, 0), pady=10, sticky="w", ipady=5)
        self.ent_district.insert(0, "Başakşehir")

        btn_row = tk.Frame(card, bg=CARD)
        btn_row.pack(fill="x", padx=20, pady=(0, 18))
        self.btn_search = RoundedButton(
            btn_row, text="Veri Madenciliğini Başlat  🔍", command=self.start_scraping_thread,
            bg=PRIMARY, hover_bg=PRIMARY_HOVER, fg="#ffffff",
            width=300, height=44, radius=12, parent_bg=CARD
        )
        self.btn_search.pack(side="left")

        # ================= TABLO =================
        table_frame = tk.Frame(self.root, bg=BG, highlightthickness=1, highlightbackground=BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=10)

        columns = ("title", "phone", "email", "website", "address")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")

        self.tree.heading("title", text="İşletme Adı")
        self.tree.heading("phone", text="Telefon")
        self.tree.heading("email", text="E-Posta")
        self.tree.heading("website", text="İşletme Web Sitesi")
        self.tree.heading("address", text="Adres")

        self.tree.column("title", width=180)
        self.tree.column("phone", width=110)
        self.tree.column("email", width=180)
        self.tree.column("website", width=200)
        self.tree.column("address", width=240)

        self.tree.tag_configure("odd", background="#0f1524")
        self.tree.tag_configure("even", background=CARD)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ================= DURUM ÇUBUĞU =================
        status_frame = tk.Frame(self.root, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        status_frame.pack(fill="x", padx=24, pady=(0, 10))

        self.lbl_status = tk.Label(status_frame, text="🛰️ Sistem Durumu: Yapay Zeka Filtre İstasyonu Aktif.",
                                    bg=CARD, fg=MUTED, font=F_STATUS, anchor="w", justify="left")
        self.lbl_status.pack(fill="x", padx=16, pady=10)

        # ================= DIŞA AKTARIM =================
        export_row = tk.Frame(self.root, bg=BG)
        export_row.pack(fill="x", padx=24, pady=(0, 22))
        self.btn_export = RoundedButton(
            export_row, text="📥  Elde Edilen Verileri Excel Olarak Dışa Aktar", command=self.export_data,
            bg=SUCCESS, hover_bg=SUCCESS_HOVER, fg="#ffffff",
            width=1, height=46, radius=12, parent_bg=BG
        )
        self.btn_export.pack(fill="x")
        export_row.update_idletasks()
        self.root.bind("<Configure>", self._resize_export_button)

    def _resize_export_button(self, _event=None):
        # Dışa aktarım butonunu pencere genişliğine göre yeniden çizer (tam genişlik, yuvarlak köşeli görünüm).
        try:
            new_width = self.btn_export.master.winfo_width()
            if new_width > 10 and new_width != self.btn_export.width:
                self.btn_export.width = new_width
                self.btn_export.config(width=new_width)
                self.btn_export.delete("all")
                pts = _round_rect_points(1, 1, new_width - 1, self.btn_export.height - 1, 12)
                self.btn_export.shape = self.btn_export.create_polygon(
                    pts, fill=self.btn_export.bg_color if not self.btn_export.disabled else self.btn_export.disabled_bg,
                    outline="", smooth=True)
                self.btn_export.label = self.btn_export.create_text(
                    new_width / 2, self.btn_export.height / 2,
                    text="📥  Elde Edilen Verileri Excel Olarak Dışa Aktar",
                    fill=self.btn_export.fg_color if not self.btn_export.disabled else self.btn_export.disabled_fg,
                    font=("Segoe UI", 10, "bold"))
        except Exception:
            pass

    def check_industry_with_ai(self, title, website, target_industry):
        # NOT: Bu fonksiyon SADECE sektör uygunluğunu (evet/hayır) belirler.
        # Mail adresi tahmini/üretimi tamamen kaldırıldı - mail artık gerçek
        # web sitesinden (find_email_on_website) taranarak bulunuyor.
        if not self.groq_client:
            return True

        prompt = f"""
        Hedef Sektör: {target_industry}
        İnceleme İstenecek İşletme Adı: {title}
        Web Sitesi: {website}

        Görevin: Bu işletmenin ismini ve varsa web adresini analiz et. Bu işletme doğrudan veya dolaylı olarak "{target_industry}" sektörü ile mi ilgileniyor yoksa tamamen alakasız başka bir iş kolu mu (büfe, restoran, alakasız tamirci vb.)?

        Yanıtı SADECE ve SADECE şu JSON formatında ver, başka hiçbir açıklama yazma:
        {{"uygun_mu": true/false}}
        """
        try:
            chat_completion = self.groq_client.chat.completions.create(
                messages=[{"role": "user", "content": prompt}],
                model="llama3-8b-8192",
                response_format={"type": "json_object"},
                timeout=5
            )
            res = json.loads(chat_completion.choices[0].message.content)
            return res.get("uygun_mu", True)
        except:
            return True

    def extract_email_from_page(self, page, pattern):
        """Açık olan bir Playwright sayfasından gerçek mail adresini çıkarır.
        Önce mailto: linklerine bakar (en güvenilir kaynak), bulamazsa
        sayfa HTML içeriğinde regex ile mail adresi arar. Rastgele/uydurma
        hiçbir şey üretmez - sadece sayfada gerçekten yazan adresi döndürür."""
        try:
            mailto_el = page.locator('a[href^="mailto:"]').first
            if mailto_el.count() > 0:
                href = mailto_el.get_attribute("href")
                if href:
                    addr = href.replace("mailto:", "").split("?")[0].strip()
                    if "@" in addr:
                        return addr
        except Exception:
            pass

        try:
            content = page.content()
            match = pattern.search(content)
            if match:
                return match.group(0)
        except Exception:
            pass

        return "Yok"

    def find_email_on_website(self, context, website):
        """Google Haritalar'dan gelen gerçek web sitesine gider, ana sayfada
        mail arar; bulamazsa 'İletişim/Contact' bağlantısını tespit edip o
        sayfaya geçer ve orada tekrar arar. Hiçbir aşamada mail uydurmaz;
        bulamazsa 'Yok' döner."""
        if not website or website == "Yok":
            return "Yok"

        email_pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
        email = "Yok"
        new_page = None
        try:
            new_page = context.new_page()
            new_page.goto(website, timeout=15000, wait_until="domcontentloaded")
            new_page.wait_for_timeout(1500)

            email = self.extract_email_from_page(new_page, email_pattern)

            if email == "Yok":
                contact_keywords = ["İletişim", "iletişim", "İLETİŞİM", "Contact", "contact", "CONTACT", "Bize Ulaşın", "bize ulaşın"]
                contact_href = None
                for kw in contact_keywords:
                    try:
                        link = new_page.locator(f'a:has-text("{kw}")').first
                        if link.count() > 0:
                            href = link.get_attribute("href")
                            if href:
                                contact_href = href
                                break
                    except Exception:
                        continue

                if contact_href:
                    contact_url = urllib.parse.urljoin(new_page.url, contact_href)
                    try:
                        new_page.goto(contact_url, timeout=15000, wait_until="domcontentloaded")
                        new_page.wait_for_timeout(1500)
                        email = self.extract_email_from_page(new_page, email_pattern)
                    except Exception:
                        pass
        except Exception:
            pass
        finally:
            if new_page:
                try:
                    new_page.close()
                except Exception:
                    pass

        return email

    def start_scraping_thread(self):
        threading.Thread(target=self.run_playwright_scraper, daemon=True).start()

    def run_playwright_scraper(self):
        kw = self.ent_keyword.get().strip()
        city = self.ent_city.get().strip()
        district = self.ent_district.get().strip()
        target_limit = int(self.spin_limit.get())

        if not kw or not city:
            messagebox.showwarning("Hata", "Lütfen Hedef Sektör ve İl alanlarını doldurun!")
            return

        self.btn_search.config(state="disabled")
        self.leads_data.clear()
        for i in self.tree.get_children():
            self.tree.delete(i)

        self.lbl_status.config(text="⚡ Sistem Uyarısı: Güvenli tarayıcı motoru ayağa kaldırılıyor...", fg="#38bdf8")

        location_query = f"{district} {city} Türkiye".strip()
        search_query = f"{kw} {location_query}"
        url = f"https://www.google.com/maps/search/{urllib.parse.quote(search_query)}"

        with sync_playwright() as p:
            try:
                browser = p.chromium.launch(headless=False, channel="chrome", args=["--disable-blink-features=AutomationControlled"])
            except Exception:
                try:
                    browser = p.chromium.launch(headless=False, channel="msedge", args=["--disable-blink-features=AutomationControlled"])
                except Exception:
                    browser = p.chromium.launch(headless=False, args=["--disable-blink-features=AutomationControlled"])
            
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
            
            page = context.new_page()
            page.goto(url)
            page.wait_for_timeout(4000)

            self.lbl_status.config(text="🛰️ Harita veritabanı açıldı. İşletmeler taranıyor...", fg="#6366f1")

            scrollable_div = page.locator('div[role="feed"]')
            scraped_titles = set()
            raw_elements = []

            for _ in range(max(1, target_limit // 3)):
                if scrollable_div.count() > 0:
                    scrollable_div.evaluate('el => el.scrollBy(0, 1500)')
                    page.wait_for_timeout(1000)
                
                cards = page.locator('div[role="feed"] a[href*="/maps/place/"]').all()
                for card in cards:
                    title = card.get_attribute("aria-label")
                    href = card.get_attribute("href")
                    if title and title not in scraped_titles:
                        scraped_titles.add(title)
                        raw_elements.append({"title": title, "url": href})
                        if len(raw_elements) >= target_limit:
                            break
                if len(raw_elements) >= target_limit:
                    break

            total_found = len(raw_elements)
            self.lbl_status.config(text=f"📊 Bilgi: Toplam {total_found} işletme tespit edildi. Yapay zeka doğrulaması başlıyor...", fg="#38bdf8")

            for index, item in enumerate(raw_elements, 1):
                try:
                    title = item["title"]
                    self.lbl_status.config(text=f"⏳ İşlem Durumu: [{index} / {total_found}] - {title} analiz ediliyor...", fg="#eab308")
                    
                    page.goto(item["url"])
                    page.wait_for_timeout(2000)

                    # Web sitesi tespiti
                    website = "Yok"
                    web_el = page.locator('a[data-item-id="authority"]').first
                    if web_el.count() > 0:
                        website = web_el.get_attribute("href").strip()

                    # 🤖 Sektör Uygunluğu Kontrolü (mail ile ilgisi yok, sadece filtre)
                    is_valid = self.check_industry_with_ai(title, website, kw)
                    if not is_valid:
                        continue

                    # Telefon tespiti
                    phone = "Yok"
                    phone_el = page.locator('button[data-item-id*="phone:tel:"]').first
                    if phone_el.count() > 0:
                        phone = phone_el.get_attribute("data-item-id").replace("phone:tel:", "").strip()

                    # Adres tespiti
                    address = "Yok"
                    address_el = page.locator('button[data-item-id="address"]').first
                    if address_el.count() > 0:
                        address = address_el.get_attribute("aria-label").replace("Adres: ", "").strip()

                    # 📧 GERÇEK MAİL TESPİTİ: Sadece işletmenin gerçek web sitesine
                    # gidilip mailto: linki veya sayfa içeriğinden bulunan mail
                    # kullanılır. Hiçbir şekilde tahmin/uydurma (info@domain vb.)
                    # yapılmaz. Web sitesi yoksa veya sitede mail bulunamazsa "Yok" kalır.
                    email = "Yok"
                    if website != "Yok":
                        self.lbl_status.config(text=f"📧 İşlem Durumu: [{index} / {total_found}] - {title} web sitesinde mail aranıyor...", fg="#eab308")
                        email = self.find_email_on_website(context, website)

                    lead_item = {
                        "title": title,
                        "phone": phone,
                        "email": email,
                        "website": website,
                        "address": address
                    }
                    self.leads_data.append(lead_item)
                    row_tag = "even" if len(self.leads_data) % 2 == 0 else "odd"
                    self.tree.insert("", "end", values=(title, phone, email, website, address), tags=(row_tag,))

                except Exception:
                    continue

            browser.close()

        self.lbl_status.config(text=f"✅ Başarılı: Filtrelenmiş veri taraması bitti. {len(self.leads_data)} kurumsal lead kaydedildi.", fg="#22c55e")
        self.btn_search.config(state="normal")

    def export_data(self):
        if not self.leads_data:
            messagebox.showwarning("Uyarı", "Dışa aktarılacak kurumsal veri kaydı bulunamadı!")
            return
        
        clean_keyword = re.sub(r'[^a-zA-Z0-9çışğöüÇİŞĞÖÜ ]', '', self.ent_keyword.get().strip())
        record_count = len(self.leads_data)
        default_filename = f"{clean_keyword}_{record_count}_Kayit.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".xlsx", 
            filetypes=[("Excel Dosyası", "*.xlsx")]
        )
        
        if file_path:
            df = pd.DataFrame(self.leads_data)
            df.columns = ["İşletme Adı", "Telefon", "E-Posta", "İşletme Web Sitesi", "Adres"]
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='B2B Müşteri Portföyü')
                
                workbook = writer.book
                worksheet = writer.sheets['B2B Müşteri Portföyü']
                
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                cell_font = Font(name="Segoe UI", size=10, color="000000")
                
                thin_border_side = Side(border_style="thin", color="D9D9D9")
                cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = cell_border
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5), start=2):
                    current_fill = zebra_fill if row_idx % 2 == 0 else white_fill
                    for cell in row:
                        cell.fill = current_fill
                        cell.font = cell_font
                        cell.border = cell_border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    worksheet.column_dimensions[col_letter].width = max(max_len + 5, 12)
                
                worksheet.row_dimensions[1].height = 26

            messagebox.showinfo("Başarılı", f"Rapor kurumsal tasarımla kaydedildi. 🚀")

if __name__ == "__main__":
    root = tk.Tk()
    app = LeadScoutEnterprise(root)
    root.mainloop()